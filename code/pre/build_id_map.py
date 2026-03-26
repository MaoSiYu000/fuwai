# -*- coding: utf-8 -*-
"""
构建“学生统一ID映射表”：
- 统一主键：XH（学号/账号，形如 pjxxxx）
- 从 data/data2、data/data3 的 xlsx 中抽取可能的学生标识字段（XSBH、SID、USERNUM、cardId、KS_XH 等）
- 基于“同值匹配 + 规范化后同值匹配”做自动对齐
- 输出：
  - output/id_map/id_map.csv：逐条映射记录（source_table/source_field/source_id -> XH）
  - output/id_map/summary.txt：覆盖率与命中情况统计

注意：部分字段（如 cardId、USERNUM）可能不是学号体系，若无法对齐会保留为未匹配。
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd

try:
    import openpyxl  # type: ignore
except Exception as e:  # pragma: no cover
    raise RuntimeError(
        "缺少依赖 openpyxl。请先安装：pip install openpyxl"
    ) from e


ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = ROOT / "data"
DATA2_DIR = DATA_DIR / "data2"
DATA3_DIR = DATA_DIR / "data3"

OUT_DIR = ROOT / "output" / "id_map"
OUT_DIR.mkdir(parents=True, exist_ok=True)

PRE_DIR = ROOT / "output" / "pre"
STU_BASIC_PRE = PRE_DIR / "学生基本信息_pre.csv"

# 为了避免被超大表拖死：对每个字段限制扫描行数与唯一ID数量
MAX_ROWS_PER_FIELD = 300_000
MAX_UNIQUE_PER_FIELD = 200_000


def _norm_id(x: object) -> str:
    """把各种单元格值规范化为可匹配的字符串ID。"""
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    s = str(x).strip()
    if not s:
        return ""
    # 去掉 Excel/CSV 里常见的 .0
    if s.endswith(".0") and s.replace(".", "", 1).isdigit():
        s = s[:-2]
    # 去引号
    if (s.startswith('"') and s.endswith('"')) or (s.startswith("'") and s.endswith("'")):
        s = s[1:-1].strip()
    return s


def _looks_like_xh(s: str) -> bool:
    """粗略判断是否像学号/账号（你们当前数据里多为 pj 开头）。"""
    if not s:
        return False
    s2 = s.lower()
    return s2.startswith("pj") and len(s2) >= 6


@dataclass(frozen=True)
class SourceField:
    file_rel: str
    sheet: str
    field: str


def _read_first_sheet_header_and_index(path: Path) -> tuple[str, dict[str, int]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = wb.sheetnames[0]
        ws = wb[sheet]
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header = [_norm_id(c) for c in row]
        idx = {h: i for i, h in enumerate(header) if h}
        return sheet, idx
    finally:
        wb.close()


def _iter_column_values(path: Path, sheet: str, col_idx: int, max_rows: int | None = None) -> Iterable[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        n = 0
        for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
            if max_rows is not None and n >= max_rows:
                break
            n += 1
            if col_idx >= len(row):
                continue
            v = _norm_id(row[col_idx])
            if v:
                yield v
    finally:
        wb.close()


def _collect_unique_ids_limited(
    path: Path,
    sheet: str,
    col_idx: int,
    max_rows: int,
    max_unique: int,
) -> tuple[set[str], bool]:
    """
    收集某列的唯一ID（去空），并在达到 max_rows 或 max_unique 时提前停止。
    返回：(seen_set, truncated)
    """
    seen: set[str] = set()
    truncated = False
    n_rows = 0
    for v in _iter_column_values(path, sheet, col_idx, max_rows=max_rows):
        n_rows += 1
        seen.add(v)
        if len(seen) >= max_unique:
            truncated = True
            break
    # 如果正好扫到了 max_rows 上限，也视为截断（可能还有更多）
    if n_rows >= max_rows:
        truncated = True
    return seen, truncated


def main() -> None:
    if not STU_BASIC_PRE.exists():
        raise FileNotFoundError(
            f"未找到学生主表：{STU_BASIC_PRE}。请先运行 pre_学生基本信息.py 生成 output/pre/学生基本信息_pre.csv"
        )

    stu = pd.read_csv(STU_BASIC_PRE, encoding="utf-8-sig")
    if "XH" not in stu.columns:
        raise ValueError("学生基本信息_pre.csv 缺少列 XH")

    xh_set = set(_norm_id(v) for v in stu["XH"].tolist())
    xh_set = {v for v in xh_set if v}

    # 可匹配的“候选学号字符串集合”：既包含原始 XH，也包含小写形态（容错）
    xh_lower_to_xh = {v.lower(): v for v in xh_set}

    # 需要扫描的文件与字段（根据你们目前新表的表头）
    targets: list[tuple[Path, list[str]]] = []
    if DATA2_DIR.exists():
        for p in sorted(DATA2_DIR.glob("*.xlsx")):
            targets.append((p, ["XSBH", "SID", "XH", "LOGIN_NAME", "CREATER_LOGIN_NAME", "USERNUM"]))
    if DATA3_DIR.exists():
        for p in sorted(DATA3_DIR.glob("*.xlsx")):
            targets.append((p, ["XSBH", "SID", "XH", "KS_XH", "USERNUM", "cardld", "cardId", "CREATER_LOGIN_NAME"]))

    rows = []
    stats = {
        "total_seen_ids": 0,
        "matched": 0,
        "unmatched": 0,
        "by_field_total": {},
        "by_field_matched": {},
    }

    for path, want_fields in targets:
        file_rel = str(path.relative_to(DATA_DIR))
        sheet, header_idx = _read_first_sheet_header_and_index(path)
        print(f"\n# 扫描 {file_rel} ({sheet})")

        for field in want_fields:
            if field not in header_idx:
                continue
            col_idx = header_idx[field]

            print(f"- 字段 {field}: 收集唯一ID中…", end="")
            seen, truncated = _collect_unique_ids_limited(
                path,
                sheet,
                col_idx,
                max_rows=MAX_ROWS_PER_FIELD,
                max_unique=MAX_UNIQUE_PER_FIELD,
            )
            print(f" unique={len(seen)}{' (truncated)' if truncated else ''}")
            if not seen:
                continue

            total_field = len(seen)
            stats["by_field_total"][f"{file_rel}:{field}"] = total_field

            matched_field = 0
            for sid in sorted(seen):
                stats["total_seen_ids"] += 1

                xh = ""
                method = ""
                confidence = ""

                # 规则 1：直接命中（大小写无关）
                sid_l = sid.lower()
                if sid in xh_set:
                    xh = sid
                    method = "exact"
                    confidence = "high"
                elif sid_l in xh_lower_to_xh:
                    xh = xh_lower_to_xh[sid_l]
                    method = "lower_exact"
                    confidence = "high"
                else:
                    # 规则 2：规范化后看起来像学号才尝试（避免把 cardId 等数字乱匹配）
                    if _looks_like_xh(sid):
                        # 有些字段可能带前后空白/引号已处理；此处不做模糊编辑距离，避免误配
                        method = "xh_like_but_not_found"
                        confidence = "low"

                if xh:
                    matched_field += 1
                    stats["matched"] += 1
                else:
                    stats["unmatched"] += 1

                rows.append(
                    {
                        "source_table": file_rel,
                        "source_sheet": sheet,
                        "source_field": field,
                        "source_id": sid,
                        "xh": xh,
                        "match_method": method,
                        "confidence": confidence,
                    }
                )

            stats["by_field_matched"][f"{file_rel}:{field}"] = matched_field

    out_df = pd.DataFrame(rows)
    out_path = OUT_DIR / "id_map.csv"
    out_df.to_csv(out_path, index=False, encoding="utf-8-sig")

    # 汇总输出
    summary_lines = []
    summary_lines.append(f"学生主表 XH 数量: {len(xh_set)}")
    summary_lines.append(f"扫描文件数: {len(targets)}")
    summary_lines.append(f"累计唯一 source_id 条数(按每表每字段去重后统计): {stats['total_seen_ids']}")
    summary_lines.append(f"匹配到 XH: {stats['matched']}")
    summary_lines.append(f"未匹配: {stats['unmatched']}")
    summary_lines.append(f"字段扫描上限: MAX_ROWS_PER_FIELD={MAX_ROWS_PER_FIELD}, MAX_UNIQUE_PER_FIELD={MAX_UNIQUE_PER_FIELD}")

    summary_lines.append("\n按 (文件:字段) 覆盖率：")
    for k in sorted(stats["by_field_total"].keys()):
        tot = stats["by_field_total"][k]
        m = stats["by_field_matched"].get(k, 0)
        rate = 0.0 if tot == 0 else (m / tot)
        summary_lines.append(f"- {k}: matched {m}/{tot} ({rate:.1%})")

    (OUT_DIR / "summary.txt").write_text("\n".join(summary_lines), encoding="utf-8-sig")

    print("== ID 映射表构建完成 ==")
    print(f"输出: {out_path}")
    print(f"摘要: {OUT_DIR / 'summary.txt'}")
    print(f"匹配到 XH: {stats['matched']} / 未匹配: {stats['unmatched']}")


if __name__ == "__main__":
    main()

